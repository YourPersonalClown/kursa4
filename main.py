import undetected_chromedriver
# import time

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from random import randint

import openpyxl
import parsing

print('working...')
book = openpyxl.open('products.xlsx', read_only=False)
sheet = book.active

driver = undetected_chromedriver.Chrome()
wait = WebDriverWait(driver, 10)

for row in range(2,sheet.max_row+1):
    prod  = sheet[row][0].value
    data = parsing.get_price_mvideo(prod, driver)

    sheet.cell(row=row, column=2, value=data[2])
    sheet.cell(row=row, column=3, value=data[3])

    data = parsing.get_price_dns(prod, driver)

    sheet.cell(row=row, column=4, value=data[2])
    sheet.cell(row=row, column=5, value=data[3])
    
book.save('products.xlsx')